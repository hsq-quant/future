from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.reporting.model_family_summary import ExperimentSpec, build_model_family_summary
from src.reporting.presentation_draft import report_notes_text
from src.reporting.report_workbook import build_report_workbook


def test_build_model_family_summary_handles_classification_and_regression(tmp_path: Path) -> None:
    classification_cv = tmp_path / "classification_cv.csv"
    classification_strategy = tmp_path / "classification_strategy.csv"
    regression_cv = tmp_path / "regression_cv.csv"
    regression_strategy = tmp_path / "regression_strategy.csv"
    variant_metrics = tmp_path / "variant_metrics.csv"

    pd.DataFrame(
        [
            {"fold_id": 1, "auc": 0.60, "accuracy": 0.55, "ic": 0.08, "threshold_used": 0.42},
            {"fold_id": 2, "auc": 0.62, "accuracy": 0.57, "ic": 0.10, "threshold_used": 0.45},
        ]
    ).to_csv(classification_cv, index=False)
    pd.DataFrame(
        [
            {
                "strategy_num_weeks": 100,
                "strategy_cumulative_return": 0.25,
                "strategy_annualized_return": 0.12,
                "strategy_annualized_volatility": 0.18,
                "strategy_sharpe_ratio": 0.67,
                "strategy_max_drawdown": -0.12,
                "benchmark_cumulative_return": 0.20,
                "cumulative_return_diff": 0.05,
                "information_ratio": 0.21,
            }
        ]
    ).to_csv(classification_strategy, index=False)
    pd.DataFrame(
        [
            {"fold_id": 1, "ic": 0.11, "mse": 0.01, "mae": 0.08, "directional_accuracy": 0.53},
            {"fold_id": 2, "ic": 0.13, "mse": 0.02, "mae": 0.07, "directional_accuracy": 0.56},
        ]
    ).to_csv(regression_cv, index=False)
    pd.DataFrame(
        [
            {
                "strategy_num_weeks": 90,
                "strategy_cumulative_return": 0.18,
                "strategy_annualized_return": 0.09,
                "strategy_annualized_volatility": 0.16,
                "strategy_sharpe_ratio": 0.56,
                "strategy_max_drawdown": -0.09,
                "benchmark_cumulative_return": 0.15,
                "cumulative_return_diff": 0.03,
                "information_ratio": 0.18,
            }
        ]
    ).to_csv(regression_strategy, index=False)
    pd.DataFrame(
        [
            {
                "variant": "long_only_factor",
                "variant_label": "Long-only factor",
                "strategy_num_weeks": 90,
                "strategy_cumulative_return": 0.22,
                "strategy_annualized_return": 0.11,
                "strategy_annualized_volatility": 0.15,
                "strategy_sharpe_ratio": 0.73,
                "strategy_max_drawdown": -0.07,
                "benchmark_cumulative_return": 0.15,
                "cumulative_return_diff": 0.07,
                "information_ratio": 0.30,
            }
        ]
    ).to_csv(variant_metrics, index=False)

    specs = [
        ExperimentSpec(
            label="V1 Chinese classification",
            family="V1",
            language_scope="zh",
            task="classification",
            execution="always_in",
            cv_metrics_path=classification_cv,
            strategy_metrics_path=classification_strategy,
        ),
        ExperimentSpec(
            label="V3 English regression",
            family="V3",
            language_scope="en",
            task="regression",
            execution="baseline_tanh",
            cv_metrics_path=regression_cv,
            strategy_metrics_path=regression_strategy,
        ),
        ExperimentSpec(
            label="V3 English regression",
            family="V3",
            language_scope="en",
            task="regression",
            execution="long_only_factor",
            cv_metrics_path=regression_cv,
            strategy_metrics_path=variant_metrics,
            strategy_variant="long_only_factor",
        ),
    ]

    summary = build_model_family_summary(specs)

    assert summary["task"].tolist() == ["classification", "regression", "regression"]
    assert summary.loc[0, "mean_auc"] == 0.61
    assert summary.loc[0, "mean_accuracy"] == 0.56
    assert summary.loc[1, "mean_ic"] == 0.12
    assert summary.loc[1, "mean_directional_accuracy"] == 0.545
    assert summary.loc[2, "execution"] == "long_only_factor"
    assert summary.loc[2, "strategy_cumulative_return"] == 0.22


def test_build_report_workbook_creates_expected_sheets(tmp_path: Path) -> None:
    output_path = tmp_path / "report.xlsx"
    summary = pd.DataFrame([{"label": "V1", "task": "classification", "value": 1}])
    classification = pd.DataFrame([{"label": "V1", "mean_auc": 0.6}])
    regression = pd.DataFrame([{"label": "V3", "mean_ic": 0.1}])
    mappings = pd.DataFrame([{"label": "V3", "execution": "long_only_factor", "sharpe": 0.7}])
    regime = pd.DataFrame([{"regime": "Recovery Bull", "strategy_cumulative_return": 0.1}])

    build_report_workbook(
        output_path,
        summary=summary,
        classification=classification,
        regression=regression,
        mappings=mappings,
        regime=regime,
        notes="line1\nline2",
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Classification", "Regression", "Mappings", "Regimes", "Notes"]
    assert workbook["Summary"]["A2"].value == "V1"
    assert workbook["Regimes"]["A2"].value == "Recovery Bull"
    assert workbook["Notes"]["A2"].value == "line1"


def test_report_notes_text_mentions_key_sheets() -> None:
    notes = report_notes_text()
    assert "01_Method" in notes
    assert "07_CumRet" in notes
    assert "11_RegimeHeat" in notes
