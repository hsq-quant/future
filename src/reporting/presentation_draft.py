from __future__ import annotations

import math
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.strategy import apply_weekly_trading_costs, summarize_strategy_vs_benchmark


ROOT = Path(__file__).resolve().parents[2]
FINAL_DIR = ROOT / "reports/final"
CHART_DIR = FINAL_DIR / "charts"

START_DATE = "2021-01-22"
END_DATE = "2025-09-05"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)

REGIMES = [
    ("Recovery Bull", "2021-01-22", "2022-02-18"),
    ("War Spike", "2022-02-25", "2022-06-24"),
    ("Post-Spike Bear", "2022-07-01", "2023-06-30"),
    ("OPEC-Supported Range", "2023-07-07", "2024-12-27"),
    ("Oversupply Bear", "2025-01-03", "2025-09-05"),
]


def _read_csv(path: str | Path) -> pd.DataFrame:
    return pd.read_csv(path)


def _read_parquet(path: str | Path) -> pd.DataFrame:
    return pd.read_parquet(path)


def _clip_weekly(frame: pd.DataFrame, *, start: str = START_DATE, end: str = END_DATE) -> pd.DataFrame:
    out = frame.copy()
    out["week_end_date"] = pd.to_datetime(out["week_end_date"]).dt.normalize()
    return out[(out["week_end_date"] >= pd.Timestamp(start)) & (out["week_end_date"] <= pd.Timestamp(end))].copy()


def paper_vs_project_table() -> pd.DataFrame:
    rows = [
        {
            "dimension": "news source",
            "paper": "Alpha Vantage News Sentiment API, energy_transportation topic",
            "ours": "Chinese mainline (V1) + English global HF mainline (V3) + 2024 supplement + RSS experiments",
        },
        {
            "dimension": "article universe",
            "paper": "29,153 deduplicated articles, 2020-01-01 to 2025-12-31",
            "ours": "Chinese and English historical news mapped to INE SC weekly horizon",
        },
        {
            "dimension": "LLM sentiment models",
            "paper": "GPT-4o and Llama 3.2-3B, plus FinBERT and Alpha Vantage baselines",
            "ours": "Qwen2.5 as the main article scorer",
        },
        {
            "dimension": "sentiment dimensions",
            "paper": "5 dimensions: relevance, polarity, intensity, uncertainty, forwardness",
            "ours": "Same 5 dimensions, same prompt rule: relevance < 0.1 => others null",
        },
        {
            "dimension": "weekly feature set",
            "paper": "31 total candidate features: GPT 11 + Llama 11 + FinBERT 5 + Alpha Vantage 4",
            "ours": "11 weekly features for a single Qwen source, then V1/V3 family comparison",
        },
        {
            "dimension": "task",
            "paper": "Binary classification only",
            "ours": "Classification + regression",
        },
        {
            "dimension": "model",
            "paper": "LightGBM binary classification + 5-fold expanding window",
            "ours": "LightGBM classification + regression + 5-fold expanding window",
        },
        {
            "dimension": "trading layer",
            "paper": "No full trading execution layer",
            "ours": "Simple classification execution + factor-style regression execution + cost/slippage scenarios",
        },
        {
            "dimension": "extra baseline difference",
            "paper": "Includes FinBERT and Alpha Vantage baselines",
            "ours": "Does not include FinBERT baseline in final comparison",
        },
    ]
    return pd.DataFrame(rows)


def feature_catalog_table() -> pd.DataFrame:
    rows = [
        ("article_count", "weekly count", "Count of scored articles in the week"),
        ("relevance_mean", "simple mean", "Average relevance score"),
        ("polarity_mean", "relevance-weighted mean", "Average directional sentiment"),
        ("intensity_mean", "relevance-weighted mean", "Average sentiment strength"),
        ("uncertainty_mean", "relevance-weighted mean", "Average uncertainty / ambiguity"),
        ("forwardness_mean", "relevance-weighted mean", "Average future orientation"),
        ("polarity_std", "std", "Within-week disagreement of polarity"),
        ("uncertainty_std", "std", "Within-week disagreement of uncertainty"),
        ("polarity_momentum", "first difference", "Change in polarity_mean vs prior week"),
        ("uncertainty_momentum", "first difference", "Change in uncertainty_mean vs prior week"),
        ("forwardness_momentum", "first difference", "Change in forwardness_mean vs prior week"),
    ]
    return pd.DataFrame(rows, columns=["feature", "aggregation", "meaning"])


def coverage_weekly_table() -> pd.DataFrame:
    zh = _read_parquet(ROOT / "data/intermediate/articles_archive_clean.parquet")
    en = _read_parquet(
        ROOT / "data/intermediate/iterations/2026-03-28-english-123-129-plus-2024-broader/articles_archive_clean_english.parquet"
    )
    weeks = _read_parquet(ROOT / "data/processed/weekly_labels.parquet")[["week_end_date"]].copy()
    weeks["week_end_date"] = pd.to_datetime(weeks["week_end_date"]).dt.normalize()
    weeks = weeks[(weeks["week_end_date"] >= pd.Timestamp("2020-01-10")) & (weeks["week_end_date"] <= pd.Timestamp(END_DATE))]

    def _counts(frame: pd.DataFrame, name: str) -> pd.DataFrame:
        out = frame.copy()
        out["week_end_date"] = pd.to_datetime(out["week_end_date"]).dt.normalize()
        out = out.groupby("week_end_date").size().rename(name).reset_index()
        return out

    merged = weeks.merge(_counts(zh, "zh_article_count"), on="week_end_date", how="left")
    merged = merged.merge(_counts(en, "en_article_count"), on="week_end_date", how="left")
    merged = merged.fillna(0)
    merged["zh_article_count"] = merged["zh_article_count"].astype(int)
    merged["en_article_count"] = merged["en_article_count"].astype(int)
    merged["total_article_count"] = merged["zh_article_count"] + merged["en_article_count"]
    return merged


def _distribution_stats(series: pd.Series) -> dict[str, float]:
    s = series.astype(float)
    return {
        "mean": float(s.mean()),
        "std": float(s.std(ddof=0)),
        "q25": float(s.quantile(0.25)),
        "median": float(s.median()),
        "q75": float(s.quantile(0.75)),
        "max": float(s.max()),
    }


def ic_distribution_tables() -> tuple[pd.DataFrame, pd.DataFrame]:
    specs = [
        ("V1 classification", ROOT / "reports/iterations/2026-03-27-single-source-baseline/cv_metrics.csv", "ic"),
        ("V3 classification", ROOT / "reports/iterations/2026-03-28-english-v3-classification-sample20/cv_metrics.csv", "ic"),
        ("V1 regression", ROOT / "reports/iterations/2026-03-28-v1-regression/cv_metrics.csv", "ic"),
        ("V3 regression", ROOT / "reports/iterations/2026-03-28-english-v3-sample20/cv_metrics.csv", "ic"),
    ]
    fold_rows: list[pd.DataFrame] = []
    stats_rows: list[dict[str, float | str]] = []
    for label, path, metric in specs:
        cv = _read_csv(path).copy()
        cv["model_family"] = label
        fold_rows.append(cv[["model_family", metric]].rename(columns={metric: "ic"}))
        s = cv[metric]
        stats_rows.append(
            {
                "model_family": label,
                "mean": float(s.mean()),
                "std": float(s.std(ddof=0)),
                "q25": float(s.quantile(0.25)),
                "median": float(s.median()),
                "q75": float(s.quantile(0.75)),
                "max": float(s.max()),
            }
        )
    return pd.concat(fold_rows, ignore_index=True), pd.DataFrame(stats_rows)


def model_family_core_table() -> pd.DataFrame:
    summary = _read_csv(FINAL_DIR / "model_family_comparison_aligned.csv")
    dedup = summary.drop_duplicates(subset=["label", "task"]).copy()
    dedup["display_name"] = dedup["label"].map(
        {
            "V1 Chinese classification": "V1 classification",
            "V3 English classification": "V3 classification",
            "V1 Chinese regression": "V1 regression",
            "V3 English regression": "V3 regression",
        }
    )
    cols = [
        "display_name",
        "task",
        "language_scope",
        "mean_ic",
        "mean_directional_accuracy",
        "mean_auc",
        "mean_accuracy",
    ]
    return dedup[cols].rename(columns={"language_scope": "language"})


def model_family_expanded_table() -> pd.DataFrame:
    specs = [
        ("V1 classification", "classification", "zh", ROOT / "reports/iterations/2026-03-27-single-source-baseline/cv_metrics.csv"),
        ("V3 classification", "classification", "en", ROOT / "reports/iterations/2026-03-28-english-v3-classification-sample20/cv_metrics.csv"),
        ("V1 regression", "regression", "zh", ROOT / "reports/iterations/2026-03-28-v1-regression/cv_metrics.csv"),
        ("V3 regression", "regression", "en", ROOT / "reports/iterations/2026-03-28-english-v3-sample20/cv_metrics.csv"),
    ]
    rows: list[dict[str, float | str]] = []
    for name, task, language, path in specs:
        cv = _read_csv(path)
        ic_stats = _distribution_stats(cv["ic"])
        row: dict[str, float | str] = {
            "display_name": name,
            "task": task,
            "language": language,
            "mean_ic": ic_stats["mean"],
            "std_ic": ic_stats["std"],
            "q25_ic": ic_stats["q25"],
            "median_ic": ic_stats["median"],
            "q75_ic": ic_stats["q75"],
            "max_ic": ic_stats["max"],
        }
        if task == "classification":
            row["mean_auc"] = float(cv["auc"].mean())
            row["mean_accuracy"] = float(cv["accuracy"].mean())
            row["max_auc"] = float(cv["auc"].max())
            row["max_accuracy"] = float(cv["accuracy"].max())
        else:
            row["mean_auc"] = np.nan
            row["mean_accuracy"] = float(cv["directional_accuracy"].mean())
            row["max_auc"] = np.nan
            row["max_accuracy"] = float(cv["directional_accuracy"].max())
        rows.append(row)
    return pd.DataFrame(rows)


def market_curve_table() -> pd.DataFrame:
    weekly = _read_parquet(ROOT / "data/processed/weekly_labels.parquet").copy()
    weekly["week_end_date"] = pd.to_datetime(weekly["week_end_date"]).dt.normalize()
    weekly = weekly[(weekly["week_end_date"] >= pd.Timestamp("2020-01-10")) & (weekly["week_end_date"] <= pd.Timestamp(END_DATE))]
    weekly["cum_log_return_pct"] = (weekly["weekly_return"].cumsum() * 100.0).astype(float)
    aligned = weekly[weekly["week_end_date"] >= pd.Timestamp(START_DATE)].copy()
    aligned["benchmark_cum_return"] = (1.0 + aligned["weekly_return"]).cumprod() - 1.0
    weekly = weekly.merge(aligned[["week_end_date", "benchmark_cum_return"]], on="week_end_date", how="left")
    return weekly


def market_daily_close_table() -> pd.DataFrame:
    daily = _read_parquet(ROOT / "data/intermediate/continuous_daily.parquet").copy()
    daily["trade_date"] = pd.to_datetime(daily["trade_date"]).dt.normalize()
    return daily[["trade_date", "close", "adjusted_close", "active_contract"]].sort_values("trade_date").reset_index(drop=True)


def cumulative_curves_table() -> pd.DataFrame:
    def _load_variant(path: Path, variant: str, series_name: str) -> pd.DataFrame:
        frame = _read_csv(path)
        frame = frame.loc[frame["variant"] == variant].copy()
        frame = _clip_weekly(frame)
        return frame[["week_end_date", "cum_return", "benchmark_cum_return"]].rename(columns={"cum_return": series_name})

    v1 = _load_variant(
        ROOT / "reports/iterations/2026-03-28-v1-classification-mapping-comparison/variant_weekly_positions.csv",
        "long_only",
        "v1_classification_long_only",
    )
    v3 = _load_variant(
        ROOT / "reports/iterations/2026-03-28-v3-english-mapping-comparison/variant_weekly_positions.csv",
        "long_only_factor",
        "v3_regression_long_only_factor",
    )
    merged = v1.merge(v3[["week_end_date", "v3_regression_long_only_factor"]], on="week_end_date", how="inner")
    merged = merged.rename(columns={"benchmark_cum_return": "benchmark_cum_return"})
    return merged


def _aligned_scenario_summary(frame: pd.DataFrame, *, scenario: str) -> dict[str, float | int]:
    window = _clip_weekly(frame)
    if scenario == "gross":
        return summarize_strategy_vs_benchmark(window)
    slippage_ticks = {"realistic_0tick": 0, "realistic_1tick": 1, "realistic_2tick": 2}[scenario]
    net = apply_weekly_trading_costs(
        window,
        commission_per_lot=40.0,
        slippage_ticks=slippage_ticks,
        tick_size=0.1,
        contract_size=1000.0,
    )
    return summarize_strategy_vs_benchmark(
        net,
        strategy_return_column="net_strategy_return",
    )


def performance_cost_table() -> pd.DataFrame:
    specs = [
        ("V1 classification", "always_in", ROOT / "reports/iterations/2026-03-28-v1-classification-mapping-comparison/variant_weekly_positions.csv"),
        ("V1 classification", "long_only", ROOT / "reports/iterations/2026-03-28-v1-classification-mapping-comparison/variant_weekly_positions.csv"),
        ("V3 classification", "always_in", ROOT / "reports/iterations/2026-03-28-v3-classification-mapping-comparison/variant_weekly_positions.csv"),
        ("V3 classification", "long_only", ROOT / "reports/iterations/2026-03-28-v3-classification-mapping-comparison/variant_weekly_positions.csv"),
        ("V1 regression", "asymmetric_long_short", ROOT / "reports/iterations/2026-03-28-v1-regression-mapping-comparison/variant_weekly_positions.csv"),
        ("V1 regression", "long_only_factor", ROOT / "reports/iterations/2026-03-28-v1-regression-mapping-comparison/variant_weekly_positions.csv"),
        ("V3 regression", "asymmetric_long_short", ROOT / "reports/iterations/2026-03-28-v3-english-mapping-comparison/variant_weekly_positions.csv"),
        ("V3 regression", "long_only_factor", ROOT / "reports/iterations/2026-03-28-v3-english-mapping-comparison/variant_weekly_positions.csv"),
    ]
    rows: list[dict[str, float | str | int]] = []
    for label, variant, path in specs:
        frame = _read_csv(path)
        frame = frame.loc[frame["variant"] == variant].copy()
        for scenario in ["gross", "realistic_0tick", "realistic_1tick", "realistic_2tick"]:
            summary = _aligned_scenario_summary(frame, scenario=scenario)
            rows.append(
                {
                    "label": label,
                    "execution": variant,
                    "scenario": scenario,
                    "strategy_cumulative_return": summary["strategy_cumulative_return"],
                    "strategy_annualized_return": summary["strategy_annualized_return"],
                    "strategy_sharpe_ratio": summary["strategy_sharpe_ratio"],
                    "strategy_max_drawdown": summary["strategy_max_drawdown"],
                    "cumulative_return_diff": summary["cumulative_return_diff"],
                    "information_ratio": summary["information_ratio"],
                }
            )
    return pd.DataFrame(rows)


def position_distribution_table() -> pd.DataFrame:
    summary = _read_csv(FINAL_DIR / "model_family_comparison_aligned.csv")
    keep = [
        ("V1 Chinese classification", "always_in"),
        ("V1 Chinese classification", "long_only"),
        ("V3 English classification", "always_in"),
        ("V3 English classification", "long_only"),
        ("V1 Chinese regression", "long_only_factor"),
        ("V3 English regression", "long_only_factor"),
    ]
    mask = pd.Series(False, index=summary.index)
    for label, execution in keep:
        mask |= ((summary["label"] == label) & (summary["execution"] == execution))
    out = summary.loc[mask, ["label", "execution", "strategy_long_share", "strategy_short_share"]].copy()
    out["flat_share"] = 1.0 - out["strategy_long_share"] - out["strategy_short_share"]
    out["display_name"] = out["label"].str.replace(" Chinese", "", regex=False).str.replace(" English", "", regex=False) + " / " + out["execution"]
    return out


def regime_selected_table() -> pd.DataFrame:
    regime = _read_csv(FINAL_DIR / "regime_comparison_aligned.csv")
    keep = [
        ("V1 Chinese classification", "long_only"),
        ("V3 English regression", "long_only_factor"),
    ]
    mask = pd.Series(False, index=regime.index)
    for label, execution in keep:
        mask |= ((regime["label"] == label) & (regime["execution"] == execution))
    selected = regime.loc[mask].copy()
    benchmark = (
        regime.groupby("regime", as_index=False)[["benchmark_cumulative_return"]]
        .first()
        .assign(label="Benchmark", execution="long_only", task="benchmark")
    )
    benchmark["window_start"] = START_DATE
    benchmark["window_end"] = END_DATE
    benchmark["strategy_num_weeks"] = np.nan
    benchmark["strategy_cumulative_return"] = benchmark["benchmark_cumulative_return"]
    benchmark["strategy_annualized_return"] = np.nan
    benchmark["strategy_sharpe_ratio"] = np.nan
    benchmark["strategy_max_drawdown"] = np.nan
    benchmark["cumulative_return_diff"] = 0.0
    benchmark["information_ratio"] = np.nan
    return pd.concat([selected, benchmark], ignore_index=True, sort=False)


def _save_chart(fig: plt.Figure, name: str) -> Path:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    path = CHART_DIR / name
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path


def chart_method_flow() -> Path:
    fig, ax = plt.subplots(figsize=(14, 7))
    ax.axis("off")
    paper_steps = ["Alpha Vantage\nenergy_transportation", "GPT-4o / Llama\n+ FinBERT baseline", "31 total features", "LightGBM\nclassification"]
    our_steps = ["Chinese + English\nhistorical sources", "Qwen2.5\n5-dim scoring", "11 features\n+ V1/V3 families", "LightGBM\nclassification + regression\n+ trading layer"]
    xs = [0.25, 0.48, 0.70, 0.91]
    for i, (x, text) in enumerate(zip(xs, paper_steps)):
        ax.text(x, 0.78, text, ha="center", va="center", bbox=dict(boxstyle="round,pad=0.5", fc="#D9EAF7", ec="#1F4E78"), fontsize=11)
        if i < len(xs) - 1:
            ax.annotate("", xy=(xs[i + 1] - 0.08, 0.78), xytext=(x + 0.08, 0.78), arrowprops=dict(arrowstyle="->", lw=2))
    for i, (x, text) in enumerate(zip(xs, our_steps)):
        ax.text(x, 0.42, text, ha="center", va="center", bbox=dict(boxstyle="round,pad=0.5", fc="#FCE4D6", ec="#C55A11"), fontsize=11)
        if i < len(xs) - 1:
            ax.annotate("", xy=(xs[i + 1] - 0.08, 0.42), xytext=(x + 0.08, 0.42), arrowprops=dict(arrowstyle="->", lw=2))
    ax.text(0.03, 0.78, "Paper", fontsize=13, fontweight="bold", ha="left", va="center")
    ax.text(0.03, 0.42, "Our project", fontsize=13, fontweight="bold", ha="left", va="center")
    ax.set_title("Paper vs Our Workflow", fontsize=16, fontweight="bold")
    return _save_chart(fig, "01_method_flow.png")


def chart_news_coverage(coverage: pd.DataFrame) -> Path:
    fig, ax = plt.subplots(figsize=(14, 5))
    zh_cap = max(1, coverage["zh_article_count"].quantile(0.97))
    en_cap = max(1, coverage["en_article_count"].quantile(0.97))
    total_cap = max(1, coverage["total_article_count"].quantile(0.97))
    zh_display = coverage["zh_article_count"].clip(upper=zh_cap)
    en_display = coverage["en_article_count"].clip(upper=en_cap)
    total_display = coverage["total_article_count"].clip(upper=total_cap)
    x = np.arange(len(coverage))
    ax.bar(x - 0.2, zh_display, width=0.4, label="Chinese", color="#4F81BD")
    ax.bar(x + 0.2, en_display, width=0.4, label="English", color="#C0504D")
    ax.plot(x, total_display, color="#2F5597", lw=1.6, label="Total")
    ax.set_title("Weekly News Coverage by Language (display clipped at P97)")
    ax.set_ylabel("Article count")
    tick_idx = np.linspace(0, len(coverage) - 1, 10, dtype=int)
    ax.set_xticks(tick_idx)
    ax.set_xticklabels(pd.to_datetime(coverage.iloc[tick_idx]["week_end_date"]).dt.strftime("%Y-%m").tolist(), rotation=45, ha="right")
    ax.legend()
    return _save_chart(fig, "03_news_coverage.png")


def chart_ic_distribution(ic_folds: pd.DataFrame, ic_stats: pd.DataFrame) -> Path:
    order = ic_stats["model_family"].tolist()
    data = [ic_folds.loc[ic_folds["model_family"] == label, "ic"].tolist() for label in order]
    fig, ax = plt.subplots(figsize=(10, 5))
    bp = ax.boxplot(data, patch_artist=True, labels=order)
    for patch in bp["boxes"]:
        patch.set(facecolor="#D9EAF7", edgecolor="#1F4E78")
    ax.scatter(range(1, len(order) + 1), ic_stats["mean"], color="red", zorder=3, label="Mean")
    ax.set_ylabel("IC")
    ax.set_title("5-fold IC Distribution by Model Family")
    ax.axhline(0, color="gray", lw=1, ls="--")
    ax.legend()
    plt.xticks(rotation=20, ha="right")
    return _save_chart(fig, "04_ic_distribution.png")


def chart_ic_acc_bars(core: pd.DataFrame) -> Path:
    labels = core["display_name"].tolist()
    metric_acc = core.apply(lambda row: row["mean_accuracy"] if row["task"] == "classification" else row["mean_directional_accuracy"], axis=1)
    metric_ic = core["mean_ic"]
    fig, axes = plt.subplots(1, 2, figsize=(12, 4.5))
    axes[0].barh(labels, metric_acc, color="#4F81BD")
    axes[0].set_title("Accuracy / Directional Accuracy")
    axes[0].set_xlim(0, max(0.65, metric_acc.max() * 1.15))
    axes[1].barh(labels, metric_ic, color="#C0504D")
    axes[1].set_title("Mean IC")
    axes[1].axvline(0, color="gray", lw=1, ls="--")
    fig.suptitle("Model Family Comparison")
    fig.tight_layout()
    return _save_chart(fig, "06_ic_acc_bars.png")


def chart_market_curve(market: pd.DataFrame) -> Path:
    fig, ax1 = plt.subplots(figsize=(14, 5))
    x = pd.to_datetime(market["week_end_date"])
    ax1.plot(x, market["weekly_close"], color="#1F4E78", label="Continuous weekly close")
    ax1.set_ylabel("Price (RMB/bbl)")
    ax2 = ax1.twinx()
    ax2.plot(x, market["cum_log_return_pct"], color="#C0504D", label="Cumulative log return (%)")
    ax2.set_ylabel("Cumulative log return (%)")
    for regime, start, end in REGIMES:
        ax1.axvspan(pd.Timestamp(start), pd.Timestamp(end), color="#D9EAF7", alpha=0.12)
    ax1.set_title("INE SC Continuous Price and Cumulative Return")
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")
    return _save_chart(fig, "07_market_curve.png")


def chart_market_daily_close(daily: pd.DataFrame) -> Path:
    fig, ax = plt.subplots(figsize=(14, 4.5))
    x = pd.to_datetime(daily["trade_date"])
    ax.plot(x, daily["adjusted_close"], color="#1F4E78", lw=1.2)
    ax.set_title("INE SC Daily Adjusted Close")
    ax.set_ylabel("Price (RMB/bbl)")
    return _save_chart(fig, "07_market_daily_close.png")


def chart_cumulative_returns(curves: pd.DataFrame) -> Path:
    fig, ax = plt.subplots(figsize=(12, 5))
    x = pd.to_datetime(curves["week_end_date"])
    ax.plot(x, curves["v1_classification_long_only"], label="V1 classification long_only", color="#4F81BD", lw=2)
    ax.plot(x, curves["v3_regression_long_only_factor"], label="V3 regression long_only_factor", color="#C0504D", lw=2)
    ax.plot(x, curves["benchmark_cum_return"], label="Benchmark long-only", color="#7F7F7F", lw=2, ls="--")
    ax.set_title("Aligned Cumulative Returns")
    ax.set_ylabel("Cumulative return")
    ax.legend()
    return _save_chart(fig, "08_cumulative_returns.png")


def chart_performance_costs(costs: pd.DataFrame) -> Path:
    subset = costs[(costs["scenario"] == "realistic_1tick") & (costs["execution"].isin(["always_in", "long_only", "long_only_factor", "asymmetric_long_short"]))].copy()
    subset["display"] = subset["label"] + " / " + subset["execution"]
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.barh(subset["display"], subset["strategy_cumulative_return"], color="#4F81BD")
    ax.axvline(0, color="gray", lw=1, ls="--")
    ax.set_title("1-tick Net Cumulative Return by Strategy")
    ax.set_xlabel("Cumulative return")
    return _save_chart(fig, "09_performance_costs.png")


def chart_position_distribution(position_df: pd.DataFrame) -> Path:
    fig, ax = plt.subplots(figsize=(12, 5))
    y = np.arange(len(position_df))
    ax.barh(y, position_df["strategy_long_share"], color="#4F81BD", label="Long")
    ax.barh(y, position_df["strategy_short_share"], left=position_df["strategy_long_share"], color="#C0504D", label="Short")
    ax.barh(
        y,
        position_df["flat_share"],
        left=position_df["strategy_long_share"] + position_df["strategy_short_share"],
        color="#A5A5A5",
        label="Flat",
    )
    ax.set_yticks(y)
    ax.set_yticklabels(position_df["display_name"])
    ax.set_xlim(0, 1)
    ax.set_title("Position Mix by Strategy")
    ax.legend()
    return _save_chart(fig, "10_position_distribution.png")


def chart_regime_bars(regime_df: pd.DataFrame) -> Path:
    pivot = regime_df.pivot_table(index="regime", columns="label", values="strategy_cumulative_return", aggfunc="first")
    order = [name for name, _, _ in REGIMES]
    pivot = pivot.reindex(order)
    fig, ax = plt.subplots(figsize=(12, 5))
    x = np.arange(len(pivot.index))
    width = 0.24
    labels = pivot.columns.tolist()
    for idx, label in enumerate(labels):
        ax.bar(x + (idx - (len(labels) - 1) / 2) * width, pivot[label], width=width, label=label)
    ax.axhline(0, color="gray", lw=1, ls="--")
    ax.set_xticks(x)
    ax.set_xticklabels(pivot.index, rotation=20, ha="right")
    ax.set_title("Regime Cumulative Return Comparison")
    ax.legend()
    return _save_chart(fig, "11_regime_bars.png")


def chart_regime_heatmap(regime_df: pd.DataFrame) -> Path:
    pivot = regime_df.pivot_table(index="label", columns="regime", values="strategy_cumulative_return", aggfunc="first")
    pivot = pivot[[name for name, _, _ in REGIMES]]
    fig, ax = plt.subplots(figsize=(12, 4.5))
    im = ax.imshow(pivot.values, cmap="RdYlGn", aspect="auto")
    ax.set_xticks(np.arange(len(pivot.columns)))
    ax.set_xticklabels(pivot.columns, rotation=20, ha="right")
    ax.set_yticks(np.arange(len(pivot.index)))
    ax.set_yticklabels(pivot.index)
    ax.set_title("Regime Return Heatmap")
    for i in range(pivot.shape[0]):
        for j in range(pivot.shape[1]):
            ax.text(j, i, f"{pivot.iloc[i, j]:.1%}", ha="center", va="center", fontsize=8)
    fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
    return _save_chart(fig, "12_regime_heatmap.png")


def _write_df(ws, df: pd.DataFrame, *, start_row: int = 1, start_col: int = 1, title: str | None = None) -> tuple[int, int]:
    row = start_row
    if title:
        ws.cell(row, start_col, title)
        ws.cell(row, start_col).font = Font(bold=True, size=14)
        row += 1
    for col_offset, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row, col_offset, str(col_name))
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r_offset, (_, data_row) in enumerate(df.iterrows(), start=row + 1):
        for c_offset, value in enumerate(data_row.tolist(), start=start_col):
            ws.cell(r_offset, c_offset, value)
    return row, row + len(df)


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 40)


def _insert_image(ws, image_path: Path, anchor: str) -> None:
    img = XLImage(str(image_path))
    img.width = img.width * 0.72
    img.height = img.height * 0.72
    ws.add_image(img, anchor)


def report_notes_text() -> str:
    return "\n".join(
        [
            "# INE SC 底稿说明",
            "",
            "- 本底稿只使用 aligned 共同窗口：2021-01-22 至 2025-09-05。",
            "- `V1 classification / long_only` 是最直观、最适合课堂展示的分类结果。",
            "- `V3 regression / long_only_factor` 是最适合解释量化因子扩展价值的结果。",
            "- `V3 classification` 模型指标更高，但最朴素的 always-in 多空执行很差，说明执行映射非常重要。",
            "- 分阶段结果请优先看 Regime 柱状图和热力图。",
            "",
            "## Workbook Sheet Guide",
            "",
            "- 01_Method: 论文与本项目的方法流程对比。",
            "- 02_Features: 11 个周频特征列表。",
            "- 03_IC_Dist: 四个模型族的 IC 折间分布。",
            "- 04_Model_Table: V1/V3、分类/回归核心指标表。",
            "- 05_IC_Acc: IC 与准确率横向对比。",
            "- 06_Market: INE SC 连续价格与累计收益曲线。",
            "- 07_CumRet: 关键策略与 benchmark 的累计收益曲线。",
            "- 08_PerfCost: 带成本/滑点的绩效对比表与图。",
            "- 09_Position: 各策略 long/short/flat 占比。",
            "- 10_RegimeBar: Regime 分阶段柱状对比。",
            "- 11_RegimeHeat: Regime 绩效热力图。",
            "- 12_Notes: 汇报口径与说明。",
            "",
        ]
    )


def build_presentation_workbook(output_path: Path) -> Path:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    method_df = paper_vs_project_table()
    feature_df = feature_catalog_table()
    ic_folds_df, ic_stats_df = ic_distribution_tables()
    model_df = model_family_core_table()
    model_expanded_df = model_family_expanded_table()
    market_df = market_curve_table()
    market_daily_df = market_daily_close_table()
    curve_df = cumulative_curves_table()
    cost_df = performance_cost_table()
    position_df = position_distribution_table()
    regime_df = regime_selected_table()
    notes = report_notes_text()

    chart_paths = {
        "method": chart_method_flow(),
        "ic_distribution": chart_ic_distribution(ic_folds_df, ic_stats_df),
        "ic_acc": chart_ic_acc_bars(model_df),
        "market": chart_market_curve(market_df),
        "market_daily": chart_market_daily_close(market_daily_df),
        "cumret": chart_cumulative_returns(curve_df),
        "cost": chart_performance_costs(cost_df),
        "position": chart_position_distribution(position_df),
        "regime_bars": chart_regime_bars(regime_df),
        "regime_heatmap": chart_regime_heatmap(regime_df),
    }

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheet_specs = [
        ("01_Method", method_df, "Paper vs Our Project", chart_paths["method"], "F2"),
        ("02_Features", feature_df, "11 Weekly Features", None, None),
        ("03_IC_Dist", ic_stats_df, "IC Distribution Summary", chart_paths["ic_distribution"], "J2"),
        ("04_Model_Table", model_expanded_df, "Model Family Comparison (expanded)", None, None),
        ("05_IC_Acc", model_df, "IC and Accuracy Comparison", chart_paths["ic_acc"], "J2"),
        ("06_Market", market_df, "INE SC Market Curve", chart_paths["market"], "J2"),
        ("07_CumRet", curve_df, "Aligned Cumulative Returns", chart_paths["cumret"], "G2"),
        ("08_PerfCost", cost_df, "Performance with Costs and Slippage", chart_paths["cost"], "K2"),
        ("09_Position", position_df, "Position Distribution", chart_paths["position"], "H2"),
        ("10_RegimeBar", regime_df, "Regime Comparison", chart_paths["regime_bars"], "J2"),
        ("11_RegimeHeat", regime_df, "Regime Heatmap Source", chart_paths["regime_heatmap"], "J2"),
    ]

    for sheet_name, df, title, image_path, anchor in sheet_specs:
        ws = wb.create_sheet(sheet_name)
        _write_df(ws, df, start_row=1, start_col=1, title=title)
        if image_path and anchor:
            _insert_image(ws, image_path, anchor)
        if sheet_name == "06_Market":
            _insert_image(ws, chart_paths["market_daily"], "J26")
        _autosize(ws)

    ws = wb.create_sheet("12_Notes")
    notes_df = pd.DataFrame({"notes": notes.splitlines()})
    _write_df(ws, notes_df, start_row=1, start_col=1, title="Report Notes")
    _autosize(ws)

    wb.save(output_path)
    return output_path
